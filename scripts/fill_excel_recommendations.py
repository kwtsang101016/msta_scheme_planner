from __future__ import annotations

import datetime as dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl


@dataclass(frozen=True)
class Recommendation:
    term: str
    reason: str


def load_prereq_map(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def build_recommendations(prereq: Dict[str, str]) -> Dict[str, Recommendation]:
    """
    Encodes the programme-level policy described in Background.txt:
    - Required: Y1T1 (fundamental).
    - Elective-Core: Y1T1 (ideally), no later than other electives.
    - Capstone/Internship: Term 2 to support summer recruitment.
    Plus course-specific sequencing to respect (where known) prereqs.
    """

    def r(term: str, reason: str) -> Recommendation:
        return Recommendation(term=term, reason=reason)

    # In the Excel we only write "Term 1/Term 2" because Year 1 and Year 2
    # students can take the same course in the same term.
    T1 = "Term 1"
    T2 = "Term 2"

    rec: Dict[str, Recommendation] = {
        # Required fundamentals
        "STA5001": r(
            T1,
            "Required and foundational; supports downstream electives (Bayesian/computational methods used widely). Take at programme start.",
        ),
        "STA5002": r(
            T1,
            "Required and foundational; establishes regression framework used by causal inference, time series, high-dimensional stats, etc. Take at programme start.",
        ),
        # Elective-Core (should be taken early)
        "STA5011": r(
            T1,
            "Elective-Core; recommended early (no later than other electives) because multivariate tools (PCA, classification, clustering) are assumed in many applied electives.",
        ),
        "STA5012": r(
            T1,
            "Elective-Core; recommended early (no later than other electives) because statistical learning underpins deep learning, NLP, AI/RL, and high-dimensional statistics.",
        ),
        "MDS5205": r(
            T1,
            "Elective-Core; recommended early (no later than other electives) so students can take other electives (finance/ML/causal) with time series modeling as a tool.",
        ),
        # Skills / enabling electives
        "MDS5111": r(
            T1,
            "Programming-enabler; recommended early to support coding-heavy electives (ML, deep learning, NLP, data analysis/visualization).",
        ),
        "MDS5117": r(
            T1,
            "Low prerequisite barrier; pairs well with Python and helps students communicate results throughout the programme.",
        ),
        "MDS5115": r(
            T1,
            "Recommended early to build data engineering/big-data workflow before project-based and applied electives.",
        ),
        "DDA5002": r(
            T1,
            "Provides optimization foundations (gradient methods) used in ML/deep learning and many modeling courses; take before ML/deep learning electives.",
        ),
        "MFE5150": r(
            T1,
            "Applied elective with broad utility; taking it early leaves room for advanced finance electives (risk/credit) later.",
        ),
        "CSC5010": r(
            T1,
            "Broad AI overview fits well early and helps students choose follow-up electives (NLP/RL/deep learning).",
        ),
        # Second-term advanced / dependent electives
        "DDA5001": r(
            T2,
            "Builds on linear algebra/probability and benefits from early-term optimization + programming; take after those foundations.",
        ),
        "MDS5122": r(
            T2,
            "Deep learning is programming- and optimization-heavy; recommended after Python and core/statistical learning exposure.",
        ),
        "AIR5066": r(
            T2,
            "Reinforcement learning is typically downstream of ML/AI basics; recommended after early-term AI/ML preparation.",
        ),
        "CSC5051": r(
            T2,
            "NLP benefits from prior ML/deep learning + programming; recommended after students have the core toolkit.",
        ),
        "STA5013": r(
            T2,
            "Causal inference relies on regression/probability foundations; recommended after required regression methods.",
        ),
        "STA5014": r(
            T2,
            "High-dimensional statistics is advanced and benefits from regression + statistical learning background; recommended after those courses.",
        ),
        "MDS5202": r(
            T2,
            "Applied regression is most valuable after required regression foundations; good follow-up consolidation course.",
        ),
        "MFE5160": r(
            T2,
            "Quant risk management is more effective after financial data analysis and solid probability/statistics preparation.",
        ),
        "MFE5190": r(
            T2,
            "Credit risk modeling is advanced; recommended after finance/data foundations (and equivalent derivatives/investment background if applicable).",
        ),
        "MBI6005": r(
            T1,
            "Offer in Term 1 to support the biomedical pathway (MBI6005 → MBI6006) and avoid stacking both in the same late term. Students can then take ML foundations (e.g., optimization/Python) early and proceed to ML/biomedical ML electives smoothly.",
        ),
        "MBI6006": r(
            T2,
            "Biomedical ML is advanced; recommended after machine learning/statistical learning + programming foundations.",
        ),
        "DDA5003": r(
            T2,
            "Stochastic processes is theory-heavy and benefits from probability foundation; recommended after fundamentals.",
        ),
        "DDA5005": r(
            T2,
            "Simulation benefits from probability/stochastic process preparation; recommended after those foundations as an applied complement.",
        ),
        "DDA6020": r(
            T1,
            "Measure-theoretic probability is highly theoretical and is a good foundation for advanced theory (e.g., advanced statistics theory). Offer in Term 1 to spread workload across years and avoid stacking multiple heavy theory courses in the same Term 2.",
        ),
        "DDA6030": r(
            T2,
            "Advanced theory should follow probability foundations; recommended after students have solid probability/statistics maturity (PhD/theory track).",
        ),
        # Capstone / internship timing policy
        "STA5020": r(
            T2,
            "Designed for Term 2 to produce portfolio-ready work and support internship/job applications for the summer immediately after Term 2.",
        ),
        "STA5021": r(
            T2,
            "Designed for Term 2 to produce portfolio-ready work and support internship/job applications for the summer immediately after Term 2.",
        ),
        "STA5022": r(
            T2,
            "Internship Training is scheduled in Term 2 to align with internship/job search and summer placement immediately after Term 2.",
        ),
    }

    # Lightweight augmentation: if prerequisites mention a course in our list,
    # nudge the course to Term 2 and add a short prereq note.
    known_codes = set(rec.keys()) | set(prereq.keys())
    for code, prereq_raw in prereq.items():
        if not prereq_raw:
            continue
        mentions = [c for c in known_codes if c != code and c in prereq_raw]
        if not mentions:
            continue
        base = rec.get(code)
        if base is None:
            continue
        if base.term == T1:
            rec[code] = Recommendation(
                term=T2,
                reason=f"{base.reason} Prerequisites mention {', '.join(sorted(set(mentions)))}; schedule after those courses where applicable.",
            )
        else:
            rec[code] = Recommendation(
                term=base.term,
                reason=f"{base.reason} Prerequisites mention {', '.join(sorted(set(mentions)))}.",
            )

    return rec


def fill_workbook(xlsx_in: Path, xlsx_out: Path, recs: Dict[str, Recommendation]) -> Tuple[int, int]:
    wb = openpyxl.load_workbook(xlsx_in)
    ws = wb[wb.sheetnames[0]]

    # Identify key columns by header name (robust to future column reorder).
    header_row = 1
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c

    code_col = headers.get("Course Code", 1)
    term_col = headers.get("Recommended Course Offering Term")
    reason_col = headers.get("Reasons")
    if term_col is None or reason_col is None:
        raise RuntimeError("Target columns not found in the Excel header row.")

    updated = 0
    skipped = 0
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not code:
            continue
        code = str(code).strip()
        rec = recs.get(code)
        if rec is None:
            skipped += 1
            continue
        ws.cell(r, term_col).value = rec.term
        ws.cell(r, reason_col).value = rec.reason
        updated += 1

    wb.save(xlsx_out)
    return updated, skipped


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="xlsx_in", default="MSTA 2026 Course List.xlsx")
    parser.add_argument("--out", dest="xlsx_out", default="")
    args = parser.parse_args()

    xlsx_in = Path(args.xlsx_in)
    if not xlsx_in.exists():
        raise RuntimeError("Excel not found: MSTA 2026 Course List.xlsx")

    prereq = load_prereq_map(Path("artifacts") / "prerequisites.json")
    recs = build_recommendations(prereq)

    if args.xlsx_out:
        xlsx_out = Path(args.xlsx_out)
    else:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
        xlsx_out = Path(f"MSTA 2026 Course List.filled_{stamp}.xlsx")
    updated, skipped = fill_workbook(xlsx_in, xlsx_out, recs)
    print(f"Wrote: {xlsx_out} (updated={updated}, no_rec={skipped})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

