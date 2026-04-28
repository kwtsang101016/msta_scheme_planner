from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


COURSE_CODE_RE = re.compile(r"\b([A-Z]{3}\d{4})\b")


def load_outlines() -> Dict[str, dict]:
    p = Path("artifacts") / "outlines.json"
    if not p.exists():
        return {}
    items = json.loads(p.read_text(encoding="utf-8"))
    return {i["course_code"]: i for i in items}


def parse_prereq_edges(course_code: str, prereq_raw: str, all_codes: List[str]) -> List[Tuple[str, str]]:
    """
    Return edges (prereq -> course_code) when prereq text mentions another
    course code in the MSTA list.
    """
    if not prereq_raw:
        return []
    edges: List[Tuple[str, str]] = []
    for c in all_codes:
        if c == course_code:
            continue
        if c in prereq_raw:
            edges.append((c, course_code))
    return edges


def main() -> int:
    # Prefer the most recently modified filled file (name sorting is unreliable, e.g. ".updated").
    xlsx_candidates = list(Path(".").glob("MSTA 2026 Course List.filled_*.xlsx"))
    if xlsx_candidates:
        xlsx_in = max(xlsx_candidates, key=lambda p: p.stat().st_mtime)
    else:
        xlsx_in = Path("MSTA 2026 Course List.xlsx")
    if not xlsx_in.exists():
        raise RuntimeError("Excel not found.")

    wb = openpyxl.load_workbook(xlsx_in)
    ws = wb[wb.sheetnames[0]]

    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c

    code_col = headers.get("Course Code", 1)
    title_col = headers.get("Course Title in ENG", 2)
    unit_col = headers.get("Unit", 4)
    type_col = headers.get("Type", 6)
    term_col = headers.get("Recommended Course Offering Term", 7)
    reason_col = headers.get("Reasons", 8)

    courses = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not code:
            continue
        code = str(code).strip()
        courses.append(
            {
                "code": code,
                "title": ws.cell(r, title_col).value or "",
                "units": ws.cell(r, unit_col).value,
                "type": ws.cell(r, type_col).value or "",
                "recommendedTerm": ws.cell(r, term_col).value or "",
                "reasons": ws.cell(r, reason_col).value or "",
            }
        )

    all_codes = [c["code"] for c in courses]
    outlines = load_outlines()

    edges: List[Tuple[str, str]] = []
    prereq_text_map: Dict[str, str] = {}
    for code in all_codes:
        o = outlines.get(code)
        prereq_raw = (o or {}).get("prerequisites_raw") or ""
        prereq_text_map[code] = prereq_raw
        edges.extend(parse_prereq_edges(code, prereq_raw, all_codes))

    out_dir = Path("webapp") / "src" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "courses.json").write_text(json.dumps(courses, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "prereq_edges.json").write_text(
        json.dumps(sorted(set(edges)), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "prereq_text.json").write_text(
        json.dumps(prereq_text_map, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Wrote {len(courses)} courses and {len(set(edges))} prereq edges into {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

